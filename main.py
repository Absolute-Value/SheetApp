import os
import sys
from PyQt6.QtWidgets import QWidget, QTableWidget, QMenuBar, QPushButton, QFileDialog, QToolBar, QComboBox, QTableWidgetItem, QApplication
from openpyxl import load_workbook, Workbook

COLUMN_NUM = 1000
ROW_NUM = 1000

class Window(QWidget):
    def __init__(self):
        super().__init__()

        self.file_path = "Book1.xlsx"
        self.setWindowTitle(os.path.basename(self.file_path))
        self.setGeometry(100, 100, 800, 600)

        self.table = QTableWidget(self)
        self.table.setGeometry(0, 60, self.width(), self.height()-60)
        self.table.cellClicked.connect(self.cell_clicked)
        self.table.cellChanged.connect(self.cell_changed)
        self.table.setColumnCount(COLUMN_NUM)
        self.table.setRowCount(ROW_NUM)

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.create_menu_bar()
        self.create_tool_bar()

    def create_menu_bar(self):
        menu_bar = QMenuBar(self)
        file_menu = menu_bar.addMenu("ファイル")

        load_action = file_menu.addAction("開く")
        load_action.triggered.connect(self.load_file)

        save_action = file_menu.addAction("保存")
        save_action.triggered.connect(self.save_file)

        save_as_action = file_menu.addAction("コピーを保存")
        save_as_action.triggered.connect(self.save_file_as)

    def create_tool_bar(self):
        self.toolbar = QToolBar(self)
        self.toolbar.setGeometry(0, 0, self.width(), 60)

        self.bold_button = QPushButton("B")
        self.bold_button.setCheckable(True)
        self.toolbar.addWidget(self.bold_button)

        self.italic_button = QPushButton("I")
        self.italic_button.setCheckable(True)
        self.toolbar.addWidget(self.italic_button)

        self.underline_button = QPushButton("U")
        self.underline_button.setCheckable(True)
        self.toolbar.addWidget(self.underline_button)

        self.font_size_combobox = QComboBox()
        self.font_size_combobox.addItems(["8", "9", "10", "11", "12", "14", "16", "18", "20", "24", "28", "32", "36", "48", "72"])
        self.toolbar.addWidget(self.font_size_combobox)

    def load_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "ファイルを開く", "", "Excelファイル (*.xlsx)")
        if file_path:
            self.load_excel_data(file_path)
            self.file_path = file_path
            self.setWindowTitle(os.path.basename(self.file_path))

    def save_file(self):
        if self.file_path == "Book1.xlsx":
            self.save_file_as()
        else:
            self.save_excel_data(self.file_path)

    def save_file_as(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "ファイルを保存", "", "Excelファイル (*.xlsx)")
        if file_path:
            self.file_path = file_path
            self.save_excel_data(self.file_path)
        
    def load_excel_data(self, file_path):
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook.active

        rd = self.sheet.row_dimensions
        for row_index in rd.keys():
            self.table.setRowHeight(row_index, int(rd[row_index].height))
        sc = self.sheet.column_dimensions
        for col_index in sc.keys():
            self.table.setColumnWidth(ord(col_index) - 65, int(sc[col_index].width*7))
            for col_index2 in range(sc[col_index].min+1, sc[col_index].max+1):
                self.table.setColumnWidth(col_index2, int(sc[col_index].width*7))

        for row_index, row in enumerate(self.sheet.iter_rows()):
            for col_index, cell in enumerate(row):
                cell_value = str(cell.value) if cell.value is not None else ""
                item = QTableWidgetItem(cell_value)
                self.table.setItem(row_index, col_index, item)
                
                font = item.font()
                font.setPointSize(int(cell.font.size))
                
                if cell.font.bold:
                    font.setBold(True)
                if cell.data_type == "n": # 数字を右寄せにする
                    item.setTextAlignment(0x0082)
                item.setFont(font)

    def save_excel_data(self, file_path):
        self.workbook.save(file_path)

    def cell_clicked(self, row, column):
        item = self.table.item(row, column)
        if item is not None:
            font = item.font()
            self.bold_button.setChecked(font.bold())
            self.italic_button.setChecked(font.italic())
            self.underline_button.setChecked(font.underline())
            self.font_size_combobox.setCurrentText(str(font.pointSize()))
        else:
            self.bold_button.setChecked(False)
            self.italic_button.setChecked(False)
            self.underline_button.setChecked(False)

    def cell_changed(self, row, column):
        item = self.table.item(row, column)
        if item is not None:
            self.sheet.cell(row=row+1, column=column+1, value=item.text())

    def resizeEvent(self, event):
        self.table.setGeometry(0, 60, self.width(), self.height()-60)
        self.toolbar.setGeometry(0, 0, self.width(), 60)
        super().resizeEvent(event)

def main():
    app = QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()